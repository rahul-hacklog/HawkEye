import face_recognition
import cv2
from openpyxl import Workbook
import datetime
from vidstab import VidStab

# Get a reference to webcam #0 (the default one)
video_capture = cv2.VideoCapture(0)
    
    
# Create a woorksheet
book=Workbook()
sheet=book.active
    
# Load images.

#image_1 = face_recognition.load_image_file("1.jpg")
#image_1_face_encoding = face_recognition.face_encodings(image_1)[0]
#face_landmarks_list = face_recognition.face_landmarks(image_1)

image_5 = face_recognition.load_image_file("ak.jpg")
image_5_face_encoding = face_recognition.face_encodings(image_5)[0]
    
#image_7 = face_recognition.load_image_file("7.jpg")
#image_7_face_encoding = face_recognition.face_encodings(image_7)[0]
    
image_3 = face_recognition.load_image_file("vi.jpg")
image_3_face_encoding = face_recognition.face_encodings(image_3)[0]
    
#image_4 = face_recognition.load_image_file("4.jpg")
#image_4_face_encoding = face_recognition.face_encodings(image_4)[0]
#import os
#strg=[]
#list = os.listdir("/home/krishna/Desktop/fa1/images") # dir is your directory path
#number_files = len(list)
#print (number_files)
#for i in range(len(list)):
#    s=list[i]
#    s=s[:-4]
#    strg.append(s)
#print(strg)

#bb=os.listdir("/home/krishna/Desktop/fa1/images")
#
#st = st[:-1]


# Create arrays of known face encodings and their names
known_face_encodings = [
#        
#        image_1_face_encoding,
        image_5_face_encoding,
#        image_7_face_encoding,
        image_3_face_encoding,
#        image_4_face_encoding
        
    ]
known_face_names = [
        
#        "1",
        "5",
#        "7",
        "3",
#        "4",
       
    ]
known_face_star=["timmy","akhil","jerry","ra","zenky","hehe"]
# Initialize some variables
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True
    
# Load present date and time
now= datetime.datetime.now()
today=now.day
month=now.month
stabilizer = VidStab(kp_method='FAST', threshold=42, nonmaxSuppression=True)
while True:
 # Grab a single frame of video
    ret, frame = video_capture.read()
    stabilized_frame = stabilizer.stabilize_frame(input_frame=frame, border_size=50)
    # Resize frame of video to 1/4 size for faster face recognition processing
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    
    # Convert the image from BGR color (which OpenCV uses) to RGB color (which face_recognition uses)
    rgb_small_frame = small_frame[:, :, ::-1]
    
    # Only process every other frame of video to save time
    if process_this_frame:
        # Find all the faces and face encodings in the current frame of video
        face_locations = face_recognition.face_locations(rgb_small_frame,model="cnn")
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)
    
    face_names = []
    for face_encoding in face_encodings:
        # See if the face is a match for the known face(s)
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
        name = "Unknown"
     
                
         # If a match was found in known_face_encodings, just use the first one.
        if True in matches:
            first_match_index = matches.index(True)
            name = known_face_names[first_match_index]
            
            print(name)
            x=0
            if int(name) in range(10):
                    sheet.cell(row=int(name)+3, column=1).value = "nan"
                    sheet.cell(row=1, column=1).value="name of the criminal"
                    sheet.cell(row=1, column=2).value = "location of the criminal on "+str(today)+"-"+str(month)+"-19"
            while x<=10:
                if sheet.cell(row=x+2, column=1).value !="nan":
                    sheet.cell(row=int(name)+2, column=1).value=known_face_star[int(name)]
                    sheet.cell(row=int(name)+2, column=2).value="found at SRM University AP"
                    x=x+1
                else:
                    break
                
                    
           
#                sheet.cell(row=2, column=1).value="absent"
    
        face_names.append(name)
    
    process_this_frame = not process_this_frame
    
    
    # Display the results
    for (top, right, bottom, left), name in zip(face_locations, face_names):
           # Scale back up face locations since the frame we detected in was scaled to 1/4 size
           top *= 4
           right *= 4
           bottom *= 4
           left *= 4
           cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
           cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
           font = cv2.FONT_HERSHEY_DUPLEX
           cv2.putText(frame, name, (left + 6, bottom - 6), font, 1.0, (255, 255, 255), 1)
    
           # Draw a label with a name below the face
    
    
    # Display the resulting image
    cv2.imshow('Video', frame)
        
    # Save Woorksheet as present month
    book.save('Don.xlsx')
    
    
    # Hit 'q' on the keyboard to quit!
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break
    
# Release handle to the webcam
video_capture.release()
cv2.destroyAllWindows()
    